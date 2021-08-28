from openpyxl import load_workbook
from Grab1stTable import FirFindLink
from Grab2ndTable import SecFindLink


def CreateTable():
    wb = load_workbook('MurdersBirthdays.xlsx')
    ws = wb.active
    ws['A1'].value = 'Murder'
    ws['B1'].value = 'Birthday'


    TableUrl = 'https://en.m.wikipedia.org/wiki/List_of_serial_killers_by_number_of_victims'

    ws = FirFindLink(ws, TableUrl)

    TableUrl = [
        'https://ru.m.wikipedia.org/wiki/%D0%9A%D0%B0%D1%82%D0%B5%D0%B3%D0%BE%D1%80%D0%B8%D1%8F:%D0%A1%D0%B5%D1%80%D0%B8%D0%B9%D0%BD%D1%8B%D0%B5_%D1%83%D0%B1%D0%B8%D0%B9%D1%86%D1%8B_%D0%BF%D0%BE_%D0%B0%D0%BB%D1%84%D0%B0%D0%B2%D0%B8%D1%82%D1%83',
        'https://ru.m.wikipedia.org/w/index.php?title=%D0%9A%D0%B0%D1%82%D0%B5%D0%B3%D0%BE%D1%80%D0%B8%D1%8F:%D0%A1%D0%B5%D1%80%D0%B8%D0%B9%D0%BD%D1%8B%D0%B5_%D1%83%D0%B1%D0%B8%D0%B9%D1%86%D1%8B_%D0%BF%D0%BE_%D0%B0%D0%BB%D1%84%D0%B0%D0%B2%D0%B8%D1%82%D1%83&pageuntil=%D0%A3%D0%BE%D1%82%D1%82%D1%81%2C+%D0%9A%D0%B0%D1%80%D0%BB#mw-pages',
        'https://ru.m.wikipedia.org/w/index.php?title=%D0%9A%D0%B0%D1%82%D0%B5%D0%B3%D0%BE%D1%80%D0%B8%D1%8F:%D0%A1%D0%B5%D1%80%D0%B8%D0%B9%D0%BD%D1%8B%D0%B5_%D1%83%D0%B1%D0%B8%D0%B9%D1%86%D1%8B_%D0%BF%D0%BE_%D0%B0%D0%BB%D1%84%D0%B0%D0%B2%D0%B8%D1%82%D1%83&pagefrom=%D0%A3%D0%BE%D1%82%D1%82%D1%81%2C+%D0%9A%D0%B0%D1%80%D0%BB#mw-pages'
    ]

    for i in TableUrl:
        ws = SecFindLink(ws, i)
    wb.save('MurdersBirthdays.xlsx')
